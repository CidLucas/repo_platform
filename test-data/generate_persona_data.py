#!/usr/bin/env python3
"""Gerador de dados de teste para as 3 personas do Blu.

Uso:
    python test-data/generate_persona_data.py

Gera planilhas CSV de notas fiscais, planilhas XLSX, documentos diversos
e arquivos de ruído para cada persona, seguindo fielmente os descritivos
em test-data/personas/*/historico-do-negocio.md

Requisitos: pip install openpyxl faker python-dateutil
"""

import csv
import io
import json
import os
import random
import uuid
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime, date, timedelta
from typing import Optional

# ─── Tentar importar openpyxl (opcional — sem ele, pula planilhas .xlsx) ───
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("[aviso] openpyxl não instalado — planilhas .xlsx não serão geradas.")
    print("         pip install openpyxl")

try:
    from faker import Faker
    fake = Faker('pt_BR')
    HAS_FAKER = True
except ImportError:
    HAS_FAKER = False
    print("[aviso] faker não instalado — usando nomes hardcoded.")
    print("         pip install faker")

random.seed(42)
if HAS_FAKER:
    Faker.seed(42)

# ═══════════════════════════════════════════════════════════════════════════
# CONSTANTES GLOBAIS
# ═══════════════════════════════════════════════════════════════════════════

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PERSONAS_DIR = os.path.join(BASE_DIR, "personas")

TODAY = date(2026, 6, 24)
YEAR_START = date(2025, 1, 1)


# ═══════════════════════════════════════════════════════════════════════════
# UTILITÁRIOS
# ═══════════════════════════════════════════════════════════════════════════

def random_date(start: date, end: date) -> date:
    """Data aleatória entre start e end."""
    delta = (end - start).days
    return start + timedelta(days=random.randint(0, delta))


def format_cpf(cpf: str) -> str:
    d = ''.join(c for c in cpf if c.isdigit())
    return f"{d[:3]}.{d[3:6]}.{d[6:9]}-{d[9:11]}"


def format_cnpj(cnpj: str) -> str:
    d = ''.join(c for c in cnpj if c.isdigit())
    return f"{d[:2]}.{d[2:5]}.{d[5:8]}/{d[8:12]}-{d[12:14]}"


def format_phone(phone: str) -> str:
    d = ''.join(c for c in phone if c.isdigit())
    if len(d) == 11:
        return f"({d[:2]}) {d[2:7]}-{d[7:]}"
    return f"({d[:2]}) {d[2:6]}-{d[6:]}"


def ensure_dir(path: str):
    os.makedirs(path, exist_ok=True)


def pick(items, n=1):
    """Escolhe n itens de uma lista."""
    if isinstance(items, list):
        return random.sample(items, min(n, len(items)))
    return random.sample(list(items), min(n, len(items)))


def ler_descritivo(persona_slug: str) -> str:
    """Lê o arquivo de descritivo da persona."""
    path = os.path.join(PERSONAS_DIR, persona_slug, "historico-do-negocio.md")
    if os.path.exists(path):
        with open(path) as f:
            return f.read()
    return ""


def write_csv(path: str, rows: list[dict]):
    """Escreve uma lista de dicionários como CSV."""
    if not rows:
        return
    ensure_dir(os.path.dirname(path))
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()), delimiter=";")
        writer.writeheader()
        writer.writerows(rows)
    print(f"  ✓ CSV: {path} ({len(rows)} linhas)")


def write_text(path: str, content: str):
    """Escreve um arquivo de texto."""
    ensure_dir(os.path.dirname(path))
    with open(path, "w", encoding="utf-8") as f:
        f.write(content)
    print(f"  ✓ TXT: {path}")


def write_xlsx(path: str, sheets: dict[str, list[list]]):
    """Escreve um arquivo XLSX com uma ou mais abas."""
    if not HAS_OPENPYXL:
        return
    ensure_dir(os.path.dirname(path))
    wb = Workbook()
    # Remove aba default
    wb.remove(wb.active)
    for sheet_name, rows in sheets.items():
        ws = wb.create_sheet(title=sheet_name[:31])
        if rows:
            # Cabeçalho em negrito
            for col_idx, val in enumerate(rows[0], 1):
                cell = ws.cell(row=1, column=col_idx, value=val)
                cell.font = Font(bold=True)
            # Dados
            for row_idx, row in enumerate(rows[1:], 2):
                for col_idx, val in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=val)
            # Auto-ajustar largura (aproximado)
            for col_idx, val in enumerate(rows[0], 1):
                max_len = max(len(str(v)) for v in [val] + [r[col_idx-1] if col_idx-1 < len(r) else "" for r in rows[1:]]
                              if v is not None)
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)
    wb.save(path)
    print(f"  ✓ XLSX: {path}")


# ═══════════════════════════════════════════════════════════════════════════
# DADOS BASE — PESSOAS FÍSICAS REUTILIZÁVEIS
# ═══════════════════════════════════════════════════════════════════════════

# Clientes que aparecem em mais de uma persona
CLIENTES_COMPARTILHADOS = {
    "Construtora Novo Norte": {
        "tipo": "PJ",
        "doc": "22.456.789/0001-10",
        "cidade": "São Paulo, SP",
    },
    "Colégio São Miguel": {
        "tipo": "PJ",
        "doc": "11.234.567/0001-90",
        "cidade": "São Paulo, SP",
    },
    "Buffet Lúcia's Food": {
        "tipo": "PJ",
        "doc": "48.792.305/0001-81",
        "cidade": "São Paulo, SP",
    },
    "Carolina Mendes": {
        "tipo": "PF",
        "doc": "384.029.170-50",
        "cidade": "São Paulo, SP",
    },
}


# ═══════════════════════════════════════════════════════════════════════════
# PERSONA 1: CAROLINA MENDES — DESIGNER AUTÔNOMA
# ═══════════════════════════════════════════════════════════════════════════

def gerar_carolina():
    slug = "carolina-design"
    print(f"\n{'='*60}")
    print(f"Persona: Carolina Mendes — Designer Autônoma")
    print(f"{'='*60}")

    base = os.path.join(PERSONAS_DIR, slug)

    # ── Clientes ──
    clientes = [
        {"nome": "Padaria Vitória", "tipo": "PJ", "doc": "12.345.678/0001-00",
         "contato": "Ana", "email": "ana@padariavitoria.com.br"},
        {"nome": "TechNova Sistemas Ltda.", "tipo": "PJ", "doc": "33.444.555/0001-66",
         "contato": "Procurement", "email": "compras@technova.com.br"},
        {"nome": "Estúdio Pilates Corpo Livre", "tipo": "PJ", "doc": "45.678.901/0001-23",
         "contato": "Carla", "email": "carla@corpolivre.com"},
        {"nome": "Empório da Vila", "tipo": "PJ", "doc": "56.789.012/0001-34",
         "contato": "Marcos", "email": "marcos@emporiodavila.com.br"},
        {"nome": "Clínica Vet Saúde Animal", "tipo": "PJ", "doc": "67.890.123/0001-45",
         "contato": "Dr. Hélio", "email": "helio@vetsaude.com"},
        {"nome": "Buffet Lúcia's Food", "tipo": "PJ", "doc": "48.792.305/0001-81",
         "contato": "Lúcia", "email": "lucia@luciasfood.com.br"},
        {"nome": "NovaTech Soluções em TI", "tipo": "PJ", "doc": "32.718.694/0001-07",
         "contato": "João", "email": "joao@novatechti.com.br"},
        # Clientes avulsos
        {"nome": "Ana Clara Barbosa", "tipo": "PF", "doc": "529.837.460-12"},
        {"nome": "Mercearia Santa Luzia", "tipo": "PJ", "doc": "78.901.234/0001-56"},
        {"nome": "Dr. Ricardo Barbosa (odontologia)", "tipo": "PJ", "doc": "89.012.345/0001-67"},
    ]

    # ── SERVIÇOS PRESTADOS (NFs de venda) ──
    servicos_descricao = [
        "Criação de identidade visual completa (logo, cores, tipografia)",
        "Design de folder institucional",
        "Criação de cardápio digital",
        "Arte para banner e material gráfico",
        "Edição e tratamento de imagens",
        "Criação de posts para redes sociais (pacote 10)",
        "Manual de marca e aplicações",
        "Diagramação de apresentação corporativa",
        "Design de rótulo de produto",
        "Reformulação de identidade visual",
        "Criação de logotipo",
        "Design de cartão de visitas e papelaria",
        "Criação de site institucional (Wix)",
        "Pacote de marketing digital mensal",
        "Arte para outdoor e anúncio impresso",
    ]

    nfs = []
    nf_counter = 0

    # Projetos específicos baseados nas memórias
    projetos_especiais = [
        # TechNova — 3 parcelas em 2024 (vamos por algumas em 2025)
        {"cliente": "TechNova Sistemas Ltda.", "data": date(2025, 3, 15),
         "descricao": "Identidade visual + manual de marca (1ª parcela)", "valor": 9333.33},
        {"cliente": "TechNova Sistemas Ltda.", "data": date(2025, 4, 15),
         "descricao": "Identidade visual + manual de marca (2ª parcela)", "valor": 9333.33},
        {"cliente": "TechNova Sistemas Ltda.", "data": date(2025, 5, 15),
         "descricao": "Identidade visual + manual de marca (3ª parcela)", "valor": 9333.34},
        # Buffet Lúcia's Food
        {"cliente": "Buffet Lúcia's Food", "data": date(2025, 6, 10),
         "descricao": "Criação de logotipo + cardápio digital", "valor": 2500.00},
        # NovaTech — logo em 2023, vamos fingir que é 2025 pra ter dado
        {"cliente": "NovaTech Soluções em TI", "data": date(2025, 2, 20),
         "descricao": "Criação de logotipo e identidade visual", "valor": 3500.00},
        # Padaria Vitória — atrasa sempre, vários serviços
        {"cliente": "Padaria Vitória", "data": date(2025, 4, 5),
         "descricao": "Design de folder promocional + banner", "valor": 1800.00},
        {"cliente": "Padaria Vitória", "data": date(2025, 7, 10),
         "descricao": "Cardápio digital e arte para redes sociais", "valor": 1200.00},
        {"cliente": "Padaria Vitória", "data": date(2025, 10, 22),
         "descricao": "Arte para panfleto e banner de final de ano", "valor": 1500.00},
        {"cliente": "Padaria Vitória", "data": date(2026, 1, 15),
         "descricao": "Reformulação cardápio + posts semanais (pacote)", "valor": 2200.00},
        # Pilates — mensal recorrente
        *[
            {"cliente": "Estúdio Pilates Corpo Livre",
             "data": date(ano, mes, random.choice([1, 5, 10, 15, 20])),
             "descricao": "Gestão de redes sociais + posts (mensal)",
             "valor": 600.00}
            for ano in [2025, 2026]
            for mes in range(1, 13)
            if date(ano, mes, 1) < TODAY
            and not (ano == 2025 and mes < 3)
        ][:20],  # limitar a 20 meses
    ]

    # Achatar
    for p in projetos_especiais:
        if isinstance(p, list):
            for sub in p:
                nf_counter += 1
                nfs.append({
                    "nf_id": f"NFSE-{2025}-{nf_counter:04d}",
                    "data_emissao": sub["data"].isoformat(),
                    "tipo": "SERVIÇO",
                    "cliente": sub["cliente"],
                    "cpf_cnpj_cliente": next(
                        (c["doc"] for c in clientes if c["nome"] == sub["cliente"]), ""),
                    "descricao": sub["descricao"],
                    "valor_servico": f"{sub['valor']:.2f}",
                    "iss_retido": "N",
                    "iss_aliquota": "5.00",
                    "iss_valor": f"{sub['valor'] * 0.05:.2f}",
                    "valor_liquido": f"{sub['valor'] * 0.95:.2f}",
                })
        else:
            nf_counter += 1
            nfs.append({
                "nf_id": f"NFSE-{p['data'].year}-{nf_counter:04d}",
                "data_emissao": p["data"].isoformat(),
                "tipo": "SERVIÇO",
                "cliente": p["cliente"],
                "cpf_cnpj_cliente": next(
                    (c["doc"] for c in clientes if c["nome"] == p["cliente"]), ""),
                "descricao": p["descricao"],
                "valor_servico": f"{p['valor']:.2f}",
                "iss_retido": "N",
                "iss_aliquota": "5.00",
                "iss_valor": f"{p['valor'] * 0.05:.2f}",
                "valor_liquido": f"{p['valor'] * 0.95:.2f}",
            })

    # Serviços avulsos aleatórios
    for _ in range(12):
        cliente = random.choice(clientes)
        valor = round(random.uniform(400, 5000), 2)
        data = random_date(date(2025, 1, 1), TODAY - timedelta(days=30))
        nf_counter += 1
        nfs.append({
            "nf_id": f"NFSE-{data.year}-{nf_counter:04d}",
            "data_emissao": data.isoformat(),
            "tipo": "SERVIÇO",
            "cliente": cliente["nome"],
            "cpf_cnpj_cliente": cliente["doc"],
            "descricao": random.choice(servicos_descricao),
            "valor_servico": f"{valor:.2f}",
            "iss_retido": random.choice(["N", "S"]),
            "iss_aliquota": "5.00",
            "iss_valor": f"{valor * 0.05:.2f}",
            "valor_liquido": f"{valor * 0.95:.2f}",
        })

    write_csv(os.path.join(base, "notas-fiscais", "nfs_servicos_prestados.csv"), nfs)

    # ── NFs de COMPRA (despesas) ──
    compras = []
    compra_counter = 0

    fornecedores_compra = [
        {"nome": "Adobe Inc.", "doc": "EXTERIOR"},
        {"nome": "iStock by Getty Images", "doc": "EXTERIOR"},
        {"nome": "Gráfica Nova Era Ltda.", "doc": "23.456.789/0001-12"},
        {"nome": "Mercado Livre (Magalu)", "doc": "03.789.654/0001-09"},
        {"nome": "Google Brasil Internet Ltda.", "doc": "06.190.287/0001-17"},
        {"nome": "Stone Pagamentos S.A.", "doc": "18.938.072/0001-82"},
        {"nome": "Transportadora Flash Ltda.", "doc": "34.567.890/0001-23"},
        {"nome": "Kalunga Comercial", "doc": "43.276.987/0001-01"},
    ]

    # Assinaturas mensais
    assinaturas = [
        {"forn": "Adobe Inc.", "desc": "Assinatura Creative Cloud (PS+AI+ID)",
         "valor": 401.00, "dia": 15},
        {"forn": "Google Brasil Internet Ltda.", "desc": "Google Workspace Individual",
         "valor": 28.00, "dia": 5},
    ]
    for a in assinaturas:
        for mes in range(1, 13):
            for ano in [2025, 2026]:
                d = date(ano, mes, a["dia"])
                if d >= TODAY:
                    continue
                compra_counter += 1
                compras.append({
                    "nf_id": f"NFCE-{ano}-{compra_counter:04d}",
                    "data_emissao": d.isoformat(),
                    "tipo": "ASSINATURA" if a["forn"] == "Adobe Inc." else "SERVIÇO",
                    "fornecedor": a["forn"],
                    "cpf_cnpj_fornecedor": a["forn"]["doc"] if isinstance(a["forn"], dict) and "doc" in a["forn"] else
                        next((f["doc"] for f in fornecedores_compra if f["nome"] == a["forn"]), ""),
                    "descricao": a["desc"],
                    "valor": f"{a['valor']:.2f}",
                })

    # Compras avulsas (iStock, gráfica, etc.)
    for _ in range(25):
        forn = random.choice(fornecedores_compra)
        if forn["nome"] == "Adobe Inc." or forn["nome"] == "Google Brasil Internet Ltda.":
            continue  # já foram cobertos
        data = random_date(date(2025, 1, 1), TODAY - timedelta(days=15))
        if forn["nome"] == "iStock by Getty Images":
            valor = round(random.uniform(40, 180), 2)
            desc = random.choice([
                "Pacote de imagens (banco de fotos)",
                "Royalty-free images (lote 10)",
                "Licenciamento de vetores",
            ])
        elif forn["nome"] == "Gráfica Nova Era Ltda.":
            valor = round(random.uniform(200, 1200), 2)
            desc = random.choice([
                "Impressão de folders (500 un.)",
                "Impressão de cartões de visita (1000 un.)",
                "Impressão de banners (3 un.)",
                "Impressão de panfletos (2000 un.)",
            ])
        elif forn["nome"] == "Kalunga Comercial":
            valor = round(random.uniform(50, 400), 2)
            desc = random.choice([
                "Papel sulfite e materiais de escritório",
                "Canetas, post-its e suprimentos",
                "Toner para impressora",
            ])
        else:
            valor = round(random.uniform(15, 150), 2)
            desc = "Entrega de materiais / frete"

        compra_counter += 1
        compras.append({
            "nf_id": f"NFCE-{data.year}-{compra_counter:04d}",
            "data_emissao": data.isoformat(),
            "tipo": "COMPRA" if forn["nome"] != "Transportadora Flash Ltda." else "FRETE",
            "fornecedor": forn["nome"],
            "cpf_cnpj_fornecedor": forn["doc"],
            "descricao": desc,
            "valor": f"{valor:.2f}",
        })

    write_csv(os.path.join(base, "notas-fiscais", "nfs_compras_despesas.csv"), compras)

    # ── ORÇAMENTOS (PDF simulados como TXT/CSV) ──
    orcamentos = []
    for i, c in enumerate(pick(clientes, 5), 1):
        valor = round(random.uniform(1500, 8000), 2)
        data = random_date(date(2025, 6, 1), TODAY - timedelta(days=45))
        orcamentos.append({
            "orcamento_id": f"ORC-2025-{i:03d}",
            "data": data.isoformat(),
            "cliente": c["nome"],
            "contato": c.get("contato", ""),
            "descricao": random.choice(servicos_descricao),
            "valor_orcado": f"{valor:.2f}",
            "status": random.choice(["APROVADO", "APROVADO", "APROVADO", "RECUSADO", "CANCELADO"]),
            "validade_dias": "15",
        })
    write_csv(os.path.join(base, "propostas-orcamentos", "orcamentos.csv"), orcamentos)

    # ── PLANILHA DE CONTROLE ("Planilha da Carol") ──
    if HAS_OPENPYXL:
        planilha_path = os.path.join(base, "planilhas", "planilha_controle_carol.xlsx")
        wb = Workbook()

        # Aba: Projetos
        ws = wb.active
        ws.title = "Projetos"
        headers = ["Projeto", "Cliente", "Data Início", "Data Entrega", "Valor",
                    "% Completo", "Status", "Observações"]
        ws.append(headers)
        projetos = [
            ["Logo + Cardápio - Buffet Lúcia", "Lúcia's Food", "2025-05-20", "2025-06-10",
             "R$ 2.500", "100%", "Entregue", "Pix na entrega"],
            ["Identidade Visual TechNova", "TechNova Sistemas", "2025-02-01", "2025-05-15",
             "R$ 28.000", "100%", "Entregue", "3 parcelas, procurement chato"],
            ["Reforma Identidade - Pilates", "Estúdio Pilates Corpo Livre", "2025-03-01", "2025-03-20",
             "R$ 600/mês", "100%", "Recorrente", "Mensal, sempre pontual"],
            ["Cardápio Padaria Vitória", "Padaria Vitória", "2025-04-01", "2025-04-05",
             "R$ 1.800", "100%", "Entregue", "ANA ATRASOU PAGAMENTO (30 dias)"],
            ["Logo NovaTech TI", "NovaTech Soluções", "2025-02-01", "2025-02-20",
             "R$ 3.500", "100%", "Entregue", "Arquivo .ai sumiu no crash de 2023 (ops)"],
            ["Posts Redes Vet Saúde", "Clínica Vet Saúde Animal", "2025-08-01", "2025-08-15",
             "R$ 1.200", "100%", "Entregue", ""],
            ["Cardápio Digital Empório", "Empório da Vila", "2025-09-01", "2025-09-10",
             "R$ 1.800", "100%", "Entregue", "Pix à vista"],
            ["Banners - Construtora Novo Norte", "Construtora Novo Norte", "2026-01-10", "2026-01-25",
             "R$ 4.200", "75%", "Em andamento", "Aguardando aprovação final"],
        ]
        for p in projetos:
            ws.append(p)
        for col in range(1, len(headers)+1):
            ws.cell(row=1, column=col).font = Font(bold=True)
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 45

        # Aba: Financeiro
        ws2 = wb.create_sheet("Financeiro")
        ws2.append(["Data", "Descrição", "Categoria", "Valor", "Tipo", "Status"])
        financeiro_data = [
            ["2025-06-10", "Recebido - Lúcia's Food", "Receita", 2500.00, "Entrada", "Recebido"],
            ["2025-03-20", "Recebido - Pilates (mensal)", "Receita", 600.00, "Entrada", "Recebido"],
            ["2025-04-15", "Recebido - TechNova (2/3)", "Receita", 9333.33, "Entrada", "Recebido"],
            ["2025-05-15", "Recebido - TechNova (3/3)", "Receita", 9333.34, "Entrada", "Recebido"],
            ["2025-05-31", "Recebido - Padaria Vitória", "Receita", 1800.00, "Entrada", "Recebido (atrasado)"],
            ["2025-01-15", "Assinatura Adobe CC", "Custo Fixo", -401.00, "Saída", "Pago"],
            ["2025-02-15", "Assinatura Adobe CC", "Custo Fixo", -401.00, "Saída", "Pago"],
            ["2025-03-15", "Assinatura Adobe CC", "Custo Fixo", -401.00, "Saída", "Pago"],
            # ... mais dados
        ]
        # Gerar mais financeiro
        for mes in range(1, 13):
            for ano in [2025, 2026]:
                d = date(ano, mes, 15)
                if d >= TODAY:
                    continue
                financeiro_data.append([
                    d.isoformat(), f"Assinatura Adobe CC", "Custo Fixo", -401.00, "Saída", "Pago"
                ])
        for row in financeiro_data:
            ws2.append(row)

        wb.save(planilha_path)
        print(f"  ✓ XLSX: {planilha_path}")

    # ── DOCUMENTOS RUÍDO ──
    ruido_dir = os.path.join(base, "documentos-ruido")

    # Lista de contatos telefônicos
    contatos = """LISTA DE CONTATOS — CAROLINA MENDES
=================================
NOME                    | TEL                     | OBS
Ana (Padaria Vitória)   | (11) 9 7654-3210       | Cliente, sempre atrasa mas indica
Carla (Pilates)         | (11) 9 5432-1098       | Mensal, pontual
Dr. Hélio (Vet)         | (11) 9 3210-9876       | Semestral
Lúcia (Buffet)          | (11) 9 9345-6789       | Boa amiga, indicou cliente
João (NovaTech)         | (11) 9 8888-7777       | Logo pendente (arquivo perdido)
Seu Ricardo (Contador)  | (11) 9 2345-6789       | R$ 200/mês, reunião 1x mês
Miguel (Gráfica)        | (11) 9 4567-8901       | Melhor gráfica da região
André (freela designer) | (11) 9 8901-2345       | Freela quando sobrecarrega
Marcela (fotógrafa)     | (11) 9 6789-0123       | Indica para jobs de foto
""".strip()
    write_text(os.path.join(ruido_dir, "lista_contatos.txt"), contatos)

    # Anotações soltas (rascunho)
    anotacoes = """ANOTAÇÕES GERAIS — CADERNO DE RASCUNHO
=============================================

**22/01/2026**
Precisa ligar pro Dr. Ricardo sobre declaração do MEI anual. 
Acho que misturei despesa pessoal (o iPad novo) na PJ.
Ele vai reclamar de novo.

**15/02/2026**
Cliente novo em potencial: Maria da floricultura perto de casa.
Quer identidade visual. Marcar café pra semana que vem.

**03/03/2026**
A Ana (Padaria Vitória) me chamou pra almoço — quer pedir mais
um trabalho mas tá com vergonha de pedir fiado. Haha.
Vou fazer o orçamento mas vou pedir sinal de 50% dessa vez.

**10/05/2026**
TechNova me chamou pra fazer a atualização do manual de marca.
Será que aceito? O procurement deles é terrível mas $$$.

**01/06/2026**
PRECISO ORGANIZAR AS PASTAS NO DRIVE.
Tá uma bagunça. Não acho o arquivo do logo da NovaTech
e o João fica me cobrando. Falei que perdi no crash de 2023
mas a verdade é que esqueci onde salvei.
"""
    write_text(os.path.join(ruido_dir, "anotacoes_pessoais.txt"), anotacoes)

    # Lista de equipamentos (para testar classificação de documento)
    equipamentos = """INVENTÁRIO DE EQUIPAMENTOS — CAROLINA MENDES
==============================================
Item                        | Quando comprou | Valor   | Status
MacBook Pro 14\" M1 Pro      | 2022          | R$ 18K  | Bom estado, bateria desgastada
iPad Air M2                 | 2025          | R$ 7.5K | Uso para esboços
Monitor LG 27\" 4K           | 2021          | R$ 2.5K | Perfeito
Mesa digitalizadora Wacom   | 2023          | R$ 1.2K | Substituto (perdi a anterior)
Impressora HP LaserJet      | 2020          | R$ 800  | Só imprime em preto
Cadeira ergonômica          | 2021          | R$ 1.5K | Mole, preciso trocar
Nobreak APC                 | 2022          | R$ 600  | Ok
Fones Sony WH-1000XM4       | 2022          | R$ 1.8K | Uso diário
""".strip()
    write_text(os.path.join(ruido_dir, "inventario_equipamentos.txt"), equipamentos)

    # Lista de aniversários (ruído máximo — lista pessoal)
    aniversarios = """ANIVERSÁRIOS PARA NÃO ESQUECER
=================================
Mãe (Célia)         — 12/03
Pai (Carlos)        — 28/07
Irmã (Juliana)      — 05/09
Ana (Padaria)       — 18/08
Carla (Pilates)     — 22/11
Lúcia               — 30/01
João (NovaTech)     — 14/06
Seu Ricardo (cont)  — 09/10
Primo Leo           — 02/04
""".strip()
    write_text(os.path.join(ruido_dir, "aniversarios.txt"), aniversarios)

    # XML de NFSe (amostra)
    xml_sample = """<?xml version="1.0" encoding="UTF-8"?>
<NFSe xmlns="http://www.abrasf.org.br/ABRASF/NFSe.xsd">
  <InfNFSe>
    <Identificacao>
      <Numero>2025-0001</Numero>
      <CodigoVerificacao>ABCD-1234-EFGH-5678</CodigoVerificacao>
    </Identificacao>
    <Prestador>
      <CNPJ>384.029.170-50</CNPJ>
      <InscricaoMunicipal>8.765.432-1</InscricaoMunicipal>
      <RazaoSocial>Carolina Mendes de Oliveira</RazaoSocial>
      <NomeFantasia>Carol Mendes Design</NomeFantasia>
      <Endereco>
        <TipoLogradouro>RUA</TipoLogradouro>
        <Logradouro>Artur de Azevedo</Logradouro>
        <Numero>1040</Numero>
        <Complemento>Apto 72</Complemento>
        <Bairro>Pinheiros</Bairro>
        <Cidade>3550308</Cidade>
        <UF>SP</UF>
        <CEP>05404003</CEP>
      </Endereco>
    </Prestador>
    <Tomador>
      <CNPJ>48.792.305/0001-81</CNPJ>
      <RazaoSocial>Lúcia Alves Freitas ME</RazaoSocial>
      <Endereco>
        <Logradouro>Rua Silva Teles</Logradouro>
        <Numero>332</Numero>
        <Bairro>Barra Funda</Bairro>
        <Cidade>3550308</Cidade>
        <UF>SP</UF>
        <CEP>01140020</CEP>
      </Endereco>
    </Tomador>
    <Servico>
      <Descricao>Criação de logotipo e cardápio digital</Descricao>
      <Valores>
        <ValorServico>2500.00</ValorServico>
        <ISSRetido>false</ISSRetido>
        <Aliquota>0.05</Aliquota>
        <ValorISS>125.00</ValorISS>
        <ValorLiquido>2375.00</ValorLiquido>
      </Valores>
    </Servico>
  </InfNFSe>
</NFSe>"""
    write_text(os.path.join(ruido_dir, "exemplo_nfse_lucia.xml"), xml_sample)

    print(f"  → {sum(1 for _ in os.listdir(os.path.join(base, 'notas-fiscais')) if _.endswith('.csv'))} CSVs de NF")
    print(f"  → Persona Carolina concluída.")


# ═══════════════════════════════════════════════════════════════════════════
# PERSONA 2: LÚCIA'S FOOD — BUFFET & CATERING
# ═══════════════════════════════════════════════════════════════════════════

def gerar_lucia():
    slug = "lucia-buffet"
    print(f"\n{'='*60}")
    print(f"Persona: Lúcia's Food — Buffet & Catering")
    print(f"{'='*60}")

    base = os.path.join(PERSONAS_DIR, slug)

    # ── CLIENTES ──
    clientes = [
        {"nome": "Salão Villa Flora Eventos", "tipo": "PJ", "doc": "44.555.666/0001-77",
         "contato": "Eliane"},
        {"nome": "Construtora Novo Norte Ltda.", "tipo": "PJ", "doc": "22.456.789/0001-10",
         "contato": "Fernanda (secretária)"},
        {"nome": "Colégio São Miguel", "tipo": "PJ", "doc": "11.234.567/0001-90",
         "contato": "Diretoria"},
        {"nome": "Dra. Regina Advocacia", "tipo": "PJ", "doc": "77.888.999/0001-44",
         "contato": "Dra. Regina"},
        {"nome": "Carolina Mendes (designer)", "tipo": "PF", "doc": "384.029.170-50"},
        {"nome": "Roberto & Camila (casamento)", "tipo": "PF", "doc": "401.273.658-90"},
        {"nome": "Família Gonçalves (15 anos)", "tipo": "PF", "doc": "529.847.163-02"},
        {"nome": "Empresa Tech Solutions", "tipo": "PJ", "doc": "88.999.000/0001-55"},
        {"nome": "Associação Comercial de SP", "tipo": "PJ", "doc": "99.000.111/0001-66"},
        {"nome": "Buffet Premium (concorrente)", "tipo": "PJ", "doc": "66.777.888/0001-99"},
        # NFs de venda para pessoas físicas diversas
    ]

    # PF genéricas para eventos
    pfs = []
    for _ in range(20):
        if HAS_FAKER:
            nome = fake.name()
        else:
            nome = random.choice(["Ana Santos", "Carlos Lima", "Marina Costa",
                                   "Pedro Alves", "Juliana Rocha"])
            nome += f" {random.randint(100,999)}"
        pfs.append({"nome": nome, "tipo": "PF", "doc": ''.join([str(random.randint(0,9)) for _ in range(11)])})

    # ── FORNECEDORES ──
    fornecedores = [
        {"nome": "Mercearia do Porto", "doc": "12.345.678/0001-00", "produto": "Camarão, peixes, frutos do mar"},
        {"nome": "CEAGESP (conta corrente)", "doc": "01.234.567/0001-89", "produto": "Hortifrúti"},
        {"nome": "Açougue do Zé", "doc": "23.456.789/0001-12", "produto": "Carnes bovinas e suínas"},
        {"nome": "Casa dos Frios Ipiranga", "doc": "34.567.890/0001-23", "produto": "Queijos e embutidos"},
        {"nome": "Distribuidora Bebidas SP", "doc": "45.678.901/0001-34", "produto": "Bebidas"},
        {"nome": "Embalagens Premium Ltda.", "doc": "56.789.012/0001-45", "produto": "Descartáveis"},
        {"nome": "Gás e Equipamentos Ltda.", "doc": "67.890.123/0001-56", "produto": "Gás P45"},
        {"nome": "Doces da Vó Nina ME", "doc": "78.901.234/0001-67", "produto": "Doces finos"},
        {"nome": "Atacadão Barra Funda", "doc": "89.012.345/0001-78", "produto": "Alimentos não perecíveis"},
        {"nome": "Padaria Pão Quente", "doc": "90.123.456/0001-89", "produto": "Pães artesanais"},
    ]

    # ── NFs de VENDA (serviços de buffet) ──
    nfs_venda = []
    nf_counter = 0

    # Eventos específicos das memórias
    eventos_especiais = [
        # Casamento Silva & Costa (dez/2025) — o desastre
        {"cliente": "Roberto & Camila (casamento)", "data": date(2025, 12, 7),
         "descricao": "Buffet completo casamento 200 convidados", "valor": 18500.00,
         "obs": "Compareceram 130. Prejuízo."},
        # Construtora Novo Norte — coffee breaks quinzenais
        *[
            {"cliente": "Construtora Novo Norte Ltda.",
             "data": date(ano, mes, random.choice([1, 15])),
             "descricao": "Coffee break reunião diretoria",
             "valor": round(random.uniform(800, 1200), 2), "obs": ""}
            for ano in [2025, 2026]
            for mes in range(1, 13)
            if date(ano, mes, 1) < TODAY
            and not (ano == 2025 and mes < 3)
        ][:20],
        # Colégio São Miguel — anual
        {"cliente": "Colégio São Miguel", "data": date(2025, 6, 20),
         "descricao": "Buffet festa junina",
         "valor": 4500.00, "obs": "Tradicional, todo ano"},
        {"cliente": "Colégio São Miguel", "data": date(2025, 12, 12),
         "descricao": "Buffet formatura 9º ano",
         "valor": 5800.00, "obs": ""},
        # Dra. Regina — mensal
        *[
            {"cliente": "Dra. Regina Advocacia",
             "data": date(ano, mes, random.choice([5, 10, 15])),
             "descricao": "Coffee break + almoço executivo",
             "valor": 1500.00, "obs": "Mensal"}
            for ano in [2025, 2026]
            for mes in range(1, 13)
            if date(ano, mes, 1) < TODAY
            and not (ano == 2025 and mes < 3)
        ][:12],
        # Salão Villa Flora (Eliane) — indicações
        {"cliente": "Salão Villa Flora Eventos", "data": date(2025, 11, 22),
         "descricao": "Coquetel inauguração nova ala do salão",
         "valor": 3200.00, "obs": "Comissão informal Eliane 10%"},
    ]

    for e in eventos_especiais:
        if isinstance(e, list):
            for sub in e:
                nf_counter += 1
                nfs_venda.append({
                    "nf_id": f"NFE-{sub['data'].year}-{nf_counter:04d}",
                    "data_emissao": sub["data"].isoformat(),
                    "natureza_operacao": "VENDA",
                    "cliente": sub["cliente"],
                    "cpf_cnpj": "",
                    "descricao_produto": sub["descricao"],
                    "valor_total": f"{sub['valor']:.2f}",
                    "icms_base": f"{sub['valor'] * 0.7:.2f}",
                    "icms_valor": f"{sub['valor'] * 0.7 * 0.12:.2f}",
                    "pis": f"{sub['valor'] * 0.0165:.2f}",
                    "cofins": f"{sub['valor'] * 0.076:.2f}",
                    "observacao": sub.get("obs", ""),
                })
        else:
            nf_counter += 1
            nfs_venda.append({
                "nf_id": f"NFE-{e['data'].year}-{nf_counter:04d}",
                "data_emissao": e["data"].isoformat(),
                "natureza_operacao": "VENDA",
                "cliente": e["cliente"],
                "cpf_cnpj": "",
                "descricao_produto": e["descricao"],
                "valor_total": f"{e['valor']:.2f}",
                "icms_base": f"{e['valor'] * 0.7:.2f}",
                "icms_valor": f"{e['valor'] * 0.7 * 0.12:.2f}",
                "pis": f"{e['valor'] * 0.0165:.2f}",
                "cofins": f"{e['valor'] * 0.076:.2f}",
                "observacao": e.get("obs", ""),
            })

    # Eventos aleatórios adicionais
    tipos_evento = [
        "Buffet completo casamento",
        "Coffee break corporativo",
        "Almoço executivo para empresa",
        "Coquetel de confraternização",
        "Buffet aniversário infantil (com recreação)",
        "Festas de 15 anos (jantar + coquetel)",
        "Brunch de final de semana",
        "Marmitas executivas (lote)",
        "Canapés para evento corporativo",
        "Buffet de formatura",
    ]

    for _ in range(15):
        c = random.choice(clientes + pfs)
        valor = round(random.uniform(800, 15000), 2)
        data = random_date(date(2025, 1, 1), TODAY - timedelta(days=20))
        nf_counter += 1
        nfs_venda.append({
            "nf_id": f"NFE-{data.year}-{nf_counter:04d}",
            "data_emissao": data.isoformat(),
            "natureza_operacao": "VENDA",
            "cliente": c["nome"],
            "cpf_cnpj": c.get("doc", ""),
            "descricao_produto": random.choice(tipos_evento),
            "valor_total": f"{valor:.2f}",
            "icms_base": f"{valor * 0.7:.2f}",
            "icms_valor": f"{valor * 0.7 * 0.12:.2f}",
            "pis": f"{valor * 0.0165:.2f}",
            "cofins": f"{valor * 0.076:.2f}",
            "observacao": "",
        })

    write_csv(os.path.join(base, "notas-fiscais", "nfs_vendas_servicos.csv"), nfs_venda)

    # ── NFs de COMPRA (insumos) ──
    nfs_compra = []
    compra_counter = 0

    # Compras sazonais — dezembro (casamentos = mais insumos)
    for mes in range(1, 13):
        for _ in range(random.randint(2, 5)):
            forn = random.choice(fornecedores)
            data = date(2025, mes, random.randint(1, 28))
            if data >= TODAY:
                continue
            # Sazonalidade: mais compras em nov/dez
            if mes in [11, 12] and random.random() > 0.3:
                continue  # skip some to keep balanced
            if mes in [1] and random.random() > 0.5:
                continue  # menos em janeiro
            if forn["nome"] == "Mercearia do Porto":
                valor = round(random.uniform(800, 2000), 2)
                desc = random.choice([
                    "Camarão cinza (5kg)", "Camarão rosa (3kg)",
                    "Filé de peixe fresco (8kg)", "Lagosta (2kg)",
                    "Lula fresca (4kg)", "Polvo (3kg)",
                ])
            elif forn["nome"] == "CEAGESP (conta corrente)":
                valor = round(random.uniform(350, 800), 2)
                desc = random.choice([
                    "Hortifrúti diverso (semanal)", "Verduras, legumes e frutas",
                    "Alface, rúcula, tomate, cebola",
                ])
            elif forn["nome"] == "Açougue do Zé":
                valor = round(random.uniform(600, 1500), 2)
                desc = random.choice([
                    "Patinho moído (10kg)", "Filé mignon (5kg)",
                    "Alcatra (8kg)", "Costela bovina (12kg)",
                ])
            elif forn["nome"] == "Doces da Vó Nina ME":
                valor = round(random.uniform(250, 900), 2)
                desc = random.choice([
                    "Bem-casados (200 un.)", "Doces finos sortidos",
                    "Tortas e bolos decorados (4 un.)",
                ])
            else:
                valor = round(random.uniform(100, 800), 2)
                desc = f"{forn['produto']} — compra regular"

            compra_counter += 1
            nfs_compra.append({
                "nf_id": f"NFE-{data.year}-{compra_counter:04d}",
                "data_emissao": data.isoformat(),
                "natureza_operacao": "COMPRA",
                "fornecedor": forn["nome"],
                "cpf_cnpj_fornecedor": forn["doc"],
                "descricao_produto": desc,
                "quantidade": "",
                "valor_unitario": "",
                "valor_total": f"{valor:.2f}",
                "icms_base": f"{valor * 0.7:.2f}",
                "icms_valor": f"{valor * 0.7 * 0.12:.2f}",
            })

    write_csv(os.path.join(base, "notas-fiscais", "nfs_compras_insumos.csv"), nfs_compra)

    # ── ORÇAMENTOS ──
    orcamentos = []
    for i in range(10):
        c = random.choice(clientes)
        data = random_date(date(2025, 6, 1), TODAY - timedelta(days=15))
        valor = round(random.uniform(2000, 22000), 2)
        orcamentos.append({
            "orcamento_id": f"ORC-EVENT-{i+1:03d}",
            "data": data.isoformat(),
            "cliente": c["nome"],
            "contato": c.get("contato", ""),
            "tipo_evento": random.choice([
                "Casamento", "15 anos", "Corporativo", "Aniversário", "Formatura"
            ]),
            "num_convidados": str(random.choice([50, 80, 100, 120, 150, 200])),
            "valor_estimado": f"{valor:.2f}",
            "status": random.choice(["APROVADO", "APROVADO", "RECUSADO", "EM_NEGOCIAÇÃO"]),
        })
    write_csv(os.path.join(base, "propostas-orcamentos", "orcamentos_eventos.csv"), orcamentos)

    # ── PLANILHA DE CUSTOS ──
    if HAS_OPENPYXL:
        planilha_path = os.path.join(base, "planilhas", "controle_fornecedores.xlsx")
        wb = Workbook()

        # Aba: Fornecedores
        ws = wb.active
        ws.title = "Fornecedores"
        ws.append(["Fornecedor", "Produto", "Contato", "Telefone", "Pontualidade (1-5)",
                    "Qualidade (1-5)", "Última Compra", "Obs"])
        fornecedores_data = [
            ["Mercearia do Porto", "Frutos do mar", "Seu Miguel", "(11) 9 1234-5678", "2", "5",
             "2026-06-20", "MELHOR QUALIDADE mas atrasa SEMPRE! Pedir 2 dias antes"],
            ["Açougue do Zé", "Carnes", "Zé", "(11) 9 2345-6789", "5", "4",
             "2026-06-22", "Confiável, entrega certinho"],
            ["CEAGESP", "Hortifrúti", "Barraca 45", "(11) 9 3456-7890", "4", "4",
             "2026-06-23", "Vou pessoalmente 2x semana"],
            ["Casa dos Frios Ipiranga", "Queijos importados", "Seu Paulo", "(11) 9 4567-8901", "4", "5",
             "2026-06-18", "Ótimo para eventos especiais"],
            ["Distribuidora Bebidas SP", "Bebidas", "Carlos", "(11) 9 5678-9012", "4", "4",
             "2026-06-10", "Devolve fechados, bom pra festa"],
            ["Embalagens Premium", "Descartáveis", "Maria", "(11) 9 6789-0123", "5", "3",
             "2026-06-15", "Preço subiu muito em 2026"],
            ["Doces da Vó Nina", "Doces finos", "Dona Nina", "(11) 9 7890-1234", "5", "5",
             "2026-05-30", "Só para eventos grandes, top"],
            ["Gás e Equipamentos", "Gás P45", "Seu Jorge", "(11) 9 8901-2345", "3", "4",
             "2026-06-12", "Manutenção câmara fria também"],
        ]
        for row in fornecedores_data:
            ws.append(row)

        # Aba: Eventos Realizados
        ws2 = wb.create_sheet("Eventos Realizados")
        ws2.append(["Evento", "Cliente", "Data", "Convidados", "Valor", "Custo Insumos",
                     "Custo M.O.", "Lucro", "Nota"])
        eventos_data = [
            ["Casamento Silva & Costa", "Roberto & Camila", "2025-12-07", 200,
             18500.00, 11000.00, 4500.00, 3000.00, "Desastre — só 130 foram"],
            ["Coffee Break Novo Norte", "Construtora Novo Norte", "2025-12-01", 20,
             950.00, 350.00, 200.00, 400.00, "Ótimo cliente"],
            ["Festa Junina Colégio", "Colégio São Miguel", "2025-06-20", 300,
             4500.00, 2200.00, 1200.00, 1100.00, "Tradição"],
            ["Aniversário 15 anos", "Família Gonçalves", "2025-08-15", 120,
             8200.00, 4200.00, 2500.00, 1500.00, "Bem organizado"],
            ["Coffee Break Advocacia", "Dra. Regina", "2025-09-10", 15,
             1500.00, 500.00, 300.00, 700.00, "Recorrente mensal"],
        ]
        for row in eventos_data:
            ws2.append(row)

        # Aba: Fluxo de Caixa
        ws3 = wb.create_sheet("Fluxo de Caixa")
        ws3.append(["Mês", "Receita", "Custos Fixos", "Custos Variáveis", "Freelas", "Lucro"])
        for mes_num in range(1, 13):
            mes_nome = {1: "Janeiro", 2: "Fevereiro", 3: "Março", 4: "Abril",
                        5: "Maio", 6: "Junho", 7: "Julho", 8: "Agosto",
                        9: "Setembro", 10: "Outubro", 11: "Novembro", 12: "Dezembro"}[mes_num]
            rec = round(random.uniform(18000, 50000), 2) if mes_num not in [1] else random.uniform(7000, 12000)
            cf = 10470.00
            cv = rec * 0.45
            fl = rec * 0.15
            lucro = rec - cf - cv - fl
            ws3.append([mes_nome, round(rec, 2), cf, round(cv, 2), round(fl, 2), round(lucro, 2)])

        wb.save(planilha_path)
        print(f"  ✓ XLSX: {planilha_path}")

    # ── DOCUMENTOS RUÍDO ──
    ruido_dir = os.path.join(base, "documentos-ruido")

    # Checklist de evento
    checklist = """CHECKLIST PRÉ-EVENTO — LÚCIA'S FOOD
=====================================
Evento: _________________________________
Data: ___________________________________
Clientes: _______________________________
Convidados previstos: ___________________

☐ Confirmar cardápio com cliente (7 dias antes)
☐ Comprar insumos perecíveis (3 dias antes)
☐ Comprar carnes e frios (2 dias antes)
☐ Retirar bebidas (1 dia antes ou no dia)
☐ Conferir estoque de descartáveis
☐ Verificar gás (cozinha + churrasqueira se houver)
☐ Confirmar freelancers (WhatsApp, 3 dias antes)
  ☐ Garçom 1: ___________
  ☐ Garçom 2: ___________
  ☐ Auxiliar cozinha: ____
☐ Separar uniformes e aventais
☐ Montar estação de café e bebidas
☐ Preparar mise en place (noite anterior)
☐ Fotos do evento (se autorizado pelo cliente)
☐ ► PÓS-EVENTO: Debriefing e prestação de contas

OBS:
__________________________________________
__________________________________________
"""
    write_text(os.path.join(ruido_dir, "checklist_pre_evento.txt"), checklist)

    # Contrato de prestação de serviços (modelo)
    contrato = """CONTRATO DE PRESTAÇÃO DE SERVIÇOS DE BUFFET
============================================

CONTRATANTE: ______________________________________ (CPF/CNPJ: ____________________)
CONTRATADA: Lúcia Alves Freitas ME — CNPJ: 48.792.305/0001-81

1. DO EVENTO
   Tipo: ___________________
   Data: ____/____/________
   Horário: ____:____ às ____:____
   Local: _________________________________
   Número de convidados: ______

2. DO VALOR
   Valor total: R$ __________________
   Sinal (50%): R$ __________________ (pagamento na assinatura)
   Saldo (50%): R$ __________________ (pagamento até 7 dias antes do evento)

3. DO CANCELAMENTO
   - Até 30 dias antes: devolução integral do sinal
   - Até 15 dias antes: retenção de 50% do sinal
   - Menos de 15 dias: sem devolução do sinal
   - Se o contratante reduzir o número de convidados em mais de 20%
     com menos de 7 dias de antecedência, não há redução de valor.

4. DA RESPONSABILIDADE
   A CONTRATADA se responsabiliza pela qualidade dos alimentos e serviços.
   Não se responsabiliza por objetos de valor deixados no local.
   O cardápio poderá ser alterado em caso de indisponibilidade de insumos,
   mantendo o padrão de qualidade.

5. OBSERVAÇÕES ADICIONAIS
   __________________________________________________
   __________________________________________________

São Paulo, ____ de ________________ de ________

___________________________          ___________________________
CONTRATANTE                          CONTRATADA (Lúcia Alves F.)
"""
    write_text(os.path.join(ruido_dir, "modelo_contrato_evento.txt"), contrato)

    # Lista de ingredientes
    ingredientes = """RECEITA — ARROZ DE FORNO ESPECIAL (para 50 pessoas)
====================================================

Ingredientes:
- Arroz arbóreo:          2,5 kg
- Frango desfiado:        3 kg
- Catupiry:               1,5 kg
- Mussarela:              1 kg
- Milho verde:            2 latas
- Ervilha:                1 lata
- Creme de leite:         4 cx
- Alho picado:            4 colheres (sopa)
- Cebola picada:          3 un
- Salsinha e cebolinha:   1 maço
- Sal, pimenta:          a gosto
- Queijo parmesão p/ gratinar: 200g

Preparo:
1. Cozinhar o arroz arbóreo (ponto al dente) no caldo de frango
2. Refogar cebola e alho, adicionar frango desfiado
3. Misturar frango ao arroz, adicionar creme de leite
4. Dispor em assadeira, cobrir com catupiry e mussarela
5. Gratinar em forno a 200°C por 20 minutos

Dica da Lúcia: O segredo é o caldo de frango caseiro (2 galinhas
cozidas por 3h com legumes). Não usar caldo em pó!
"""
    write_text(os.path.join(ruido_dir, "receita_arroz_forno.txt"), ingredientes)

    # Lista de equipamentos
    equipamentos = """INVENTÁRIO DE EQUIPAMENTOS — COZINHA INDUSTRIAL
=================================================
Item                          | Quant. | Aquisição | Valor   | Estado
Fogão industrial 6 bocas      | 2      | 2020      | R$ 4.5K | Bom
Fogão industrial 4 bocas      | 1      | 2018      | R$ 2.8K | Desgastado
Câmara fria                   | 1      | 2019      | R$ 12K  | Manutenção em dia
Freezer horizontal            | 2      | 2021      | R$ 3.2K | Ok
Geladeira industrial          | 1      | 2022      | R$ 4K   | Ok
Batedeira planetária          | 1      | 2020      | R$ 1.8K | Precisa revisão
Liquidificador industrial     | 2      | 2021      | R$ 800  | Um queimou em maio
Forno elétrico                | 2      | 2022      | R$ 2.5K | Bom
Exaustor                      | 1      | 2020      | R$ 3K   | Trocar filtros
Marmitex térmico transp       | 8      | 2023      | R$ 400  | Novos
Panelas alumínio (conjunto)   | 3 conj | 2019      | R$ 1.5K | Precisam ser trocadas
Facas profissionais (kit)     | 2 kits | 2021      | R$ 600  | Amolar urgente
Utenślios diversos (conjunto) | 1      | 2020      | R$ 400  | Ok
Balança digital               | 2      | 2022      | R$ 300  | Ok
Microondas industrial         | 1      | 2021      | R$ 1.2K | Ok
Lava-louças industrial        | 1      | 2019      | R$ 5K   | Com defeito desde jan/26
"""
    write_text(os.path.join(ruido_dir, "inventario_equipamentos.txt"), equipamentos)

    # Exemplo de NF XML
    xml = """<?xml version="1.0" encoding="UTF-8"?>
<NFe xmlns="http://www.portalfiscal.inf.br/nfe">
  <infNFe Id="NFe35200648792305000181550010000000011000000011" versao="4.00">
    <ide>
      <cUF>35</cUF>
      <cNF>00000001</cNF>
      <natOp>VENDA</natOp>
      <mod>55</mod>
      <serie>1</serie>
      <nNF>1</nNF>
      <dhEmi>2025-12-07T10:00:00-03:00</dhEmi>
      <tpNF>1</tpNF>
      <idDest>1</idDest>
      <tpImp>1</tpImp>
      <tpEmis>1</tpEmis>
      <cDV>1</cDV>
      <tpAmb>1</tpAmb>
      <finNFe>1</finNFe>
      <indFinal>1</indFinal>
      <indPres>1</indPres>
      <procEmi>0</procEmi>
      <verProc>teste_gerador</verProc>
    </ide>
    <emit>
      <CNPJ>48792305000181</CNPJ>
      <xNome>Lucia Alves Freitas ME</xNome>
      <xFant>Lucia's Food Buffet &amp; Catering</xFant>
      <enderEmit>
        <xLgr>Rua Silva Teles</xLgr>
        <nro>332</nro>
        <xBairro>Barra Funda</xBairro>
        <xMun>Sao Paulo</xMun>
        <UF>SP</UF>
        <CEP>01140020</CEP>
      </enderEmit>
    </emit>
    <dest>
      <CPF>40127365890</CPF>
      <xNome>Roberto Silva e Camila Costa</xNome>
    </dest>
    <det nItem="1">
      <prod>
        <cProd>001</cProd>
        <xProd>Buffet Completo Casamento - 200 convidados</xProd>
        <CFOP>5102</CFOP>
        <uCom>SERV</uCom>
        <qCom>1</qCom>
        <vUnCom>18500.00</vUnCom>
        <vProd>18500.00</vProd>
      </prod>
      <imposto>
        <vTotTrib>4320.00</vTotTrib>
        <ICMS15>
          <orig>0</orig>
          <CST>00</CST>
          <modBC>3</modBC>
          <vBC>12950.00</vBC>
          <pICMS>12.00</pICMS>
          <vICMS>1554.00</vICMS>
        </ICMS15>
      </imposto>
    </det>
    <total>
      <ICMSTot>
        <vBC>12950.00</vBC>
        <vICMS>1554.00</vICMS>
        <vPIS>305.25</vPIS>
        <vCOFINS>1406.00</vCOFINS>
        <vNF>18500.00</vNF>
      </ICMSTot>
    </total>
  </infNFe>
</NFe>"""
    write_text(os.path.join(ruido_dir, "exemplo_nfe_casamento.xml"), xml)

    # Lista de convidados (ruído)
    convidados = """LISTA DE CONVIDADOS — CASAMENTO SILVA & COSTA
================================================
ATENÇÃO: LISTA ATUALIZADA EM 01/12/2025

CONFIRMADOS (130):
01. Roberto Silva (noivo)
02. Camila Costa (noiva)
03. Sr. Antônio Silva (pai do noivo)
04. Sra. Márcia Silva (mãe do noivo)
...
[dados truncados para exemplo]

NÃO CONFIRMARAM (70):
...

OBSERVAÇÃO DA LÚCIA: Fizemos comida pra 200 e só vieram 130.
Prejuízo total de mais de R$ 8 mil.
Da próxima vez, exijo 50% de sinal.
"""
    write_text(os.path.join(ruido_dir, "lista_convidados_casamento.txt"), convidados)

    print(f"  → {sum(1 for _ in os.listdir(os.path.join(base, 'notas-fiscais')) if _.endswith('.csv'))} CSVs de NF")
    print(f"  → Persona Lúcia concluída.")


# ═══════════════════════════════════════════════════════════════════════════
# PERSONA 3: NOVATECH TI — EMPRESA DE 4 FUNCIONÁRIOS
# ═══════════════════════════════════════════════════════════════════════════

def gerar_novatech():
    slug = "novatech-ti"
    print(f"\n{'='*60}")
    print(f"Persona: NovaTech Soluções em TI")
    print(f"{'='*60}")

    base = os.path.join(PERSONAS_DIR, slug)

    # ── CLIENTES ──
    clientes = [
        {"nome": "AutoPeças Lima Ltda.", "tipo": "PJ", "doc": "11.222.333/0001-44",
         "contato": "Seu Lima", "plano": "R$ 2.500/mês"},
        {"nome": "Escritório Advocacia Rocha & Silva", "tipo": "PJ", "doc": "44.555.666/0001-77",
         "contato": "Dr. Rocha", "plano": "R$ 4.800/mês"},
        {"nome": "Farmácia Bem-Estar Ltda.", "tipo": "PJ", "doc": "77.888.999/0001-00",
         "contato": "Sr. Paulo", "plano": "R$ 3.200/mês"},
        {"nome": "Mecânica do Zé", "tipo": "PJ", "doc": "55.666.777/0001-88",
         "contato": "Zé", "plano": "R$ 600/mês"},
        {"nome": "Restaurante do Beto", "tipo": "PJ", "doc": "66.777.888/0001-99",
         "contato": "Beto", "plano": "R$ 1.500/mês"},
        {"nome": "Construtora Novo Norte Ltda.", "tipo": "PJ", "doc": "22.456.789/0001-10",
         "contato": "Eng. Carlos", "plano": "R$ 3.800/mês"},
        {"nome": "Consultório Dr. Fernando (dentista)", "tipo": "PJ", "doc": "33.444.555/0001-66",
         "contato": "Dr. Fernando", "plano": "R$ 500/mês"},
        {"nome": "Buffet Lúcia's Food", "tipo": "PJ", "doc": "48.792.305/0001-81",
         "contato": "Lúcia", "plano": "R$ 200/mês"},
        {"nome": "Colégio São Miguel", "tipo": "PJ", "doc": "11.234.567/0001-90",
         "contato": "Secretaria"},
        # Clientes de hardware avulsos
        {"nome": "Padaria Vitória", "tipo": "PJ", "doc": "12.345.678/0001-00"},
        {"nome": "Empório da Vila", "tipo": "PJ", "doc": "56.789.012/0001-34"},
    ]

    # ── FORNECEDORES ──
    fornecedores = [
        {"nome": "DigiDistri Distribuidora", "doc": "12.345.678/0001-00"},
        {"nome": "Intelbras S.A.", "doc": "83.123.456/0001-00"},
        {"nome": "Dell Computadores do Brasil", "doc": "04.123.456/0001-00"},
        {"nome": "AWS Brasil (Amazon)", "doc": "EXTERIOR"},
        {"nome": "Escritório Contábil Tatuapé (Sr. Paulo)", "doc": "91.234.567/0001-88"},
        {"nome": "Mercado Livre / Magalu", "doc": "03.789.654/0001-09"},
        {"nome": "Escritório Central de Marcas", "doc": "81.234.567/0001-77"},
    ]

    # ── NFs de VENDA (SERVIÇOS - contratos mensais) ──
    nfs_venda_serv = []
    nf_counter = 0

    # Contratos mensais recorrentes
    for cliente in clientes:
        if not cliente.get("plano"):
            continue
        valor_plano = float(cliente["plano"].replace("R$ ", "").replace("/mês", "").replace(".", "").replace(",", "."))
        for mes in range(1, 13):
            for ano in [2025, 2026]:
                d = date(ano, mes, random.randint(1, 5))
                if d >= TODAY:
                    continue
                # Cliente começou em algum momento
                if ano == 2025 and mes < {  # data de início aproximada
                    "AutoPeças Lima Ltda.": 1,
                    "Escritório Advocacia Rocha & Silva": 1,
                    "Farmácia Bem-Estar Ltda.": 4,
                    "Mecânica do Zé": 1,
                    "Restaurante do Beto": 4,
                    "Construtora Novo Norte Ltda.": 1,
                    "Consultório Dr. Fernando (dentista)": 3,
                    "Buffet Lúcia's Food": 6,
                }.get(cliente["nome"], 1):
                    continue
                nf_counter += 1
                nfs_venda_serv.append({
                    "nf_id": f"NFE-{d.year}-{nf_counter:04d}",
                    "data_emissao": d.isoformat(),
                    "natureza_operacao": "SERVICO",
                    "cliente": cliente["nome"],
                    "cpf_cnpj": cliente["doc"],
                    "descricao": f"Manutenção preventiva mensal ({cliente['plano']})",
                    "valor_total": f"{valor_plano:.2f}",
                    "iss_retido": "N",
                    "iss_aliquota": "5.00",
                    "iss_valor": f"{valor_plano * 0.05:.2f}",
                    "observacao": "",
                })

    write_csv(os.path.join(base, "notas-fiscais", "nfs_vendas_servicos_recorrentes.csv"), nfs_venda_serv)

    # ── NFs de VENDA (HARDWARE) ──
    vendas_hw = [
        # Baseado nas memórias
        {"cliente": "AutoPeças Lima Ltda.", "data": date(2026, 3, 15),
         "descricao": "10 Monitores Dell 24\" + 5 SSDs Kingston 480GB",
         "valor": 9800.00},
        {"cliente": "Escritório Advocacia Rocha & Silva", "data": date(2026, 2, 10),
         "descricao": "3 Notebooks Dell Latitude 3540",
         "valor": 12400.00},
        {"cliente": "Farmácia Bem-Estar Ltda.", "data": date(2025, 12, 5),
         "descricao": "Servidor Dell PowerEdge T160 + Nobreak APC 1500VA",
         "valor": 18500.00},
        {"cliente": "Colégio São Miguel", "data": date(2026, 1, 20),
         "descricao": "20 Chromebooks educacionais Samsung",
         "valor": 24000.00},
        # Vendas avulsas
        {"cliente": "Padaria Vitória", "data": date(2025, 8, 12),
         "descricao": "2 computadores completos + impressora térmica",
         "valor": 5200.00},
        {"cliente": "Construtora Novo Norte Ltda.", "data": date(2025, 11, 20),
         "descricao": "Switch 24 portas + cabeamento estruturado (4 andares)",
         "valor": 8500.00},
        {"cliente": "Empório da Vila", "data": date(2026, 4, 2),
         "descricao": "Leitor de código de barras + TEF + impressora não-fiscal",
         "valor": 3200.00},
        {"cliente": "Restaurante do Beto", "data": date(2025, 9, 8),
         "descricao": "2 tablets 10\" + suporte PDV",
         "valor": 2400.00},
        {"cliente": "Construtora Novo Norte Ltda.", "data": date(2026, 3, 1),
         "descricao": "Projeto migração nuvem — 30 estações (projeto completo)",
         "valor": 45000.00},
    ]

    for v in vendas_hw:
        nf_counter += 1
        nfs_venda_serv.append({
            "nf_id": f"NFE-{v['data'].year}-{nf_counter:04d}",
            "data_emissao": v['data'].isoformat(),
            "natureza_operacao": "VENDA_HW",
            "cliente": v['cliente'],
            "cpf_cnpj": next((c["doc"] for c in clientes if c["nome"] == v['cliente']), ""),
            "descricao": v['descricao'],
            "valor_total": f"{v['valor']:.2f}",
            "iss_retido": "N/A",
            "iss_aliquota": "N/A",
            "iss_valor": "0.00",
            "observacao": "",
        })

    write_csv(os.path.join(base, "notas-fiscais", "nfs_vendas_hardware.csv"),
              nfs_venda_serv[len(nfs_venda_serv)-len(vendas_hw):])

    # ── NFs de COMPRA (estoque) ──
    nfs_compra = []
    compra_counter = 0

    for _ in range(30):
        forn = random.choice(fornecedores + [
            {"nome": "DigiDistri Distribuidora", "doc": "12.345.678/0001-00"},
            {"nome": "Intelbras S.A.", "doc": "83.123.456/0001-00"},
        ])
        data = random_date(date(2025, 1, 1), TODAY - timedelta(days=10))
        valor = round(random.uniform(400, 15000), 2)

        if forn["nome"] == "DigiDistri Distribuidora":
            desc = random.choice([
                "Notebooks Dell Latitude (lote 5)",
                "Monitores Dell 24\" (lote 10)",
                "SSDs Kingston 480GB (lote 20)",
                "Teclados e mouses (kit 50)",
                "Impressoras HP LaserJet (lote 3)",
                "Cabos de rede e conectores (caixa)",
            ])
        elif forn["nome"] == "Intelbras S.A.":
            desc = random.choice([
                "Roteador empresarial (5 un.)",
                "Switch gerenciável 24 portas",
                "Câmeras IP HD (lote 8)",
                "Access Point Wi-Fi 6 (3 un.)",
                "Nobreak Intelbras 1500VA (2 un.)",
            ])
        else:
            desc = random.choice([
                "Suprimentos diversos (cabos, fontes, adaptadores)",
                "Materiais de escritório e TI",
                "Peças para reposição",
            ])

        # DigiDistri: alguns pagamentos com observação
        obs = ""
        if forn["nome"] == "DigiDistri Distribuidora" and random.random() < 0.2:
            obs = "Pagamento somente após conferência (regra Maria)"

        compra_counter += 1
        nfs_compra.append({
            "nf_id": f"NFE-{data.year}-{compra_counter:04d}",
            "data_emissao": data.isoformat(),
            "natureza_operacao": "COMPRA",
            "fornecedor": forn["nome"],
            "cpf_cnpj_fornecedor": forn["doc"],
            "descricao_produto": desc,
            "valor_total": f"{valor:.2f}",
            "icms_base": f"{valor * 0.8:.2f}",
            "icms_valor": f"{valor * 0.8 * 0.12:.2f}",
            "observacao": obs,
        })

    write_csv(os.path.join(base, "notas-fiscais", "nfs_compras_estoque.csv"), nfs_compra)

    # ── PROPOSTAS COMERCIAIS ──
    propostas = []
    for i in range(8):
        c = random.choice(clientes)
        data = random_date(date(2025, 6, 1), TODAY - timedelta(days=15))
        propostas.append({
            "proposta_id": f"PROP-{i+1:03d}",
            "data": data.isoformat(),
            "cliente": c["nome"],
            "contato": c.get("contato", ""),
            "tipo": random.choice(["MANUTENÇÃO", "HARDWARE", "PROJETO", "NUVEM"]),
            "descricao": f"Proposta de serviços de TI — {random.choice(['contrato anual', 'aquisição equipamentos', 'migração nuvem'])}",
            "valor_estimado": f"{random.uniform(1500, 35000):.2f}",
            "status": random.choice(["APROVADO", "APROVADO", "EM_NEGOCIAÇÃO", "RECUSADO"]),
        })
    write_csv(os.path.join(base, "propostas-orcamentos", "propostas_comerciais.csv"), propostas)

    # ── PLANILHAS ──
    if HAS_OPENPYXL:
        # Planilha de fluxo de caixa (Maria)
        planilha_path = os.path.join(base, "planilhas", "fluxo_caixa_2025.xlsx")
        wb = Workbook()

        ws = wb.active
        ws.title = "Fluxo de Caixa"
        ws.append(["Data", "Descrição", "Categoria", "Valor", "Saldo", "Tipo"])
        # Gerar movimentações
        saldo = 35000.00
        lancamentos = []
        # Receita contratos
        for mes in range(1, 13):
            for c in clientes:
                if not c.get("plano"):
                    continue
                valor = float(c["plano"].replace("R$ ", "").replace("/mês", "").replace(".", "").replace(",", "."))
                d = date(2025, mes, random.randint(1, 5))
                lancamentos.append((d, f"Mensalidade {c['nome']}", "Receita", valor, "Entrada"))
        # Despesas
        for mes in range(1, 13):
            # Salários
            lancamentos.append((date(2025, mes, 5), "Salários funcionários", "Custo Fixo", -15000.00, "Saída"))
            # Aluguel
            lancamentos.append((date(2025, mes, 10), "Aluguel sala Tatuapé", "Custo Fixo", -3500.00, "Saída"))
            # Contador
            lancamentos.append((date(2025, mes, 15), "Contabilidade (Sr. Paulo)", "Custo Fixo", -900.00, "Saída"))

        lancamentos.sort(key=lambda x: x[0])
        ws.append(["Data", "Descrição", "Categoria", "Valor", "Saldo", "Tipo"])
        for d, desc, cat, valor, tipo in lancamentos:
            saldo += valor
            ws.append([d.isoformat(), desc, cat, f"R$ {valor:,.2f}", f"R$ {saldo:,.2f}", tipo])

        wb.save(planilha_path)
        print(f"  ✓ XLSX: {planilha_path}")

        # Planilha de comissões
        comissoes_path = os.path.join(base, "planilhas", "comissoes_vendas.xlsx")
        wb2 = Workbook()
        ws2 = wb2.active
        ws2.title = "Comissões"
        ws2.append(["Vendedor", "Cliente", "Produto/Serviço", "Valor Venda", "% Comissão", "Valor Comissão", "Pago?"])
        comissoes_data = [
            ["João", "AutoPeças Lima", "Monitores + SSDs", 9800.00, "5%", 490.00, "Sim"],
            ["João", "Advocacia Rocha & Silva", "Notebooks Dell", 12400.00, "5%", 620.00, "Sim"],
            ["João", "Colégio São Miguel", "20 Chromebooks", 24000.00, "8%", 1920.00, "Pendente"],
            ["João", "Farmácia Bem-Estar", "Servidor Dell + Nobreak", 18500.00, "5%", 925.00, "Sim"],
            ["Rafael", "Restaurante do Beto", "Tablets + PDV", 2400.00, "3%", 72.00, "Sim"],
            ["Rafael", "Construtora Novo Norte", "Projeto migração nuvem", 45000.00, "6%", 2700.00, "Pendente"],
            ["João", "Empório da Vila", "Leitor + TEF", 3200.00, "3%", 96.00, "Pendente"],
        ]
        for row in comissoes_data:
            ws2.append(row)
        wb2.save(comissoes_path)
        print(f"  ✓ XLSX: {comissoes_path}")

    # ── DOCUMENTOS RUÍDO ──
    ruido_dir = os.path.join(base, "documentos-ruido")

    # Contrato de prestação (modelo)
    contrato = """CONTRATO DE PRESTAÇÃO DE SERVIÇOS DE TI
=========================================
CONTRATANTE: ______________________________________ (CNPJ: _____________________)
CONTRATADA: NovaTech Soluções em Informática Ltda. — CNPJ: 32.718.694/0001-07

CLÁUSULA 1 — OBJETO
Prestação de serviços de suporte técnico em TI, incluindo:
- Manutenção preventiva e corretiva de equipamentos
- Suporte remoto e presencial
- Gestão de rede e servidores
- Backup e segurança da informação

CLÁUSULA 2 — VALOR E FORMA DE PAGAMENTO
Valor mensal: R$ __________________
Pagamento: boleto bancário, todo dia 10 do mês corrente
Reajuste: anualmente pelo IPCA

CLÁUSULA 3 — PRAZO
Prazo mínimo: 12 meses
Rescisão antecipada: multa de 30% sobre saldo do contrato

CLÁUSULA 4 — ATENDIMENTO
Disponibilidade: dias úteis, 8h-18h
Emergência: WhatsApp (11) 9 8888-7777 (taxa adicional fora do horário)
Tempo de resposta: até 4h úteis

CLÁUSULA 5 — OBSERVAÇÕES
- Peças e hardware necessários para reparo serão cobrados à parte
- Visitas técnicas adicionais (fora do escopo mensal) serão cobradas à hora técnica

São Paulo, ____ de ________________ de ________

___________________________          ___________________________
CONTRATANTE                          NOVATECH TI (João Nogueira)
"""
    write_text(os.path.join(ruido_dir, "modelo_contrato_ti.txt"), contrato)

    # Nota de empenho fictícia
    empenho = """NOTA DE EMPENHO — PREFEITURA MUNICIPAL (FICTÍCIA)
====================================================
Nº Empenho: 2026/004578
Data: 15/03/2026
Credor: NovaTech Soluções em Informática Ltda.
CNPJ: 32.718.694/0001-07
Valor: R$ 18.750,00
Modalidade: Pregão Eletrônico
Processo: 2025/12.345
Descrição: Aquisição de 15 (quinze) computadores completos + suporte
           para instalação na Secretaria Municipal de Educação.
Dotação Orçamentária: 04.122.0015.2.100 — Material Permanente

OBS: Empenho aguardando dotação — previsão pagamento 60 dias.
"""
    write_text(os.path.join(ruido_dir, "nota_empenho_prefeitura.txt"), empenho)

    # Lista de senhas (ruído, péssima prática)
    senhas = """CONTROLE DE SENHAS — NOVATECH TI
==================================
ATENÇÃO: Arquivo interno. Não compartilhar.

CLIENTES:
- AutoPeças Lima: admin / Lima@2023! (servidor local)
- Rocha & Silva: suporte / Roch@Adv2024 (servidor)
- Farmácia Bem-Estar: adm / Farmac!a2024 (sistema)
- Construtora Novo Norte: adm / N0v0N0rte! (Google Workspace)

INTERNO:
- Servidor AWS: root / N0v@T3ch!2025 (MUDAR!)
- Wi-Fi escritório: NovaTechTI / 0987654321
- E-mail administrativo: admin@novatechti.com.br / NT!adm2024
- Conta azul (Maria): maria@novatechti.com.br / M@ri@2025

MARIA: Por favor, trocar senhas a cada 3 meses!!!
"""
    write_text(os.path.join(ruido_dir, "controle_senhas_interno.txt"), senhas)

    # Relatório de chamados (simulação)
    chamados = """RELATÓRIO DE CHAMADOS — MAIO 2026
=====================================

Total: 47 chamados
Abertos: 3
Em andamento: 5
Fechados: 39

TOP 5 CLIENTES:
1. Farmácia Bem-Estar — 12 chamados (sistema PDV lento)
2. Rocha & Silva — 8 chamados (e-mail fora do ar)
3. Construtora Novo Norte — 6 chamados
4. AutoPeças Lima — 5 chamados
5. Restaurante do Beto — 5 chamados (impressora vive enrolando)

CHAMADOS ABERTOS (críticos):
#1023 — Farmácia Bem-Estar — Servidor de banco fora do ar (Rafael alocado)
#1024 — Rocha & Silva — Vírus em 3 máquinas (quarentena, aguardando)

TEMPO MÉDIO DE RESPOSTA: 2,3 horas
TEMPO MÉDIO DE RESOLUÇÃO: 6,8 horas

OBSERVAÇÕES DA MARIA:
- O Rafael resolveu 90% dos chamados sozinho esse mês
- Lucas já está atendendo chamados simples (instalação, senha)
- João fechou proposta de R$ 18K na Construtora
"""
    write_text(os.path.join(ruido_dir, "relatorio_chamados_maio2026.txt"), chamados)

    # Inventário de equipamentos
    inventario = """INVENTÁRIO — ESTOQUE NOVATECH TI (ATUALIZADO EM 20/06/2026)
=================================================================

EQUIPAMENTOS DISPONÍVEIS:
Item                       | Qtd | Custo Unit. | Preço Venda
Notebook Dell Latitude     | 3   | R$ 2.200    | R$ 3.800
Notebook Samsung           | 2   | R$ 1.800    | R$ 3.200
Monitor Dell 24\"           | 5   | R$ 650      | R$ 1.100
SSD Kingston 480GB         | 12  | R$ 280      | R$ 450
SSD Kingston 240GB         | 8   | R$ 150      | R$ 260
Switch Intelbras 24p       | 1   | R$ 890      | R$ 1.490
Roteador Intelbras         | 4   | R$ 180      | R$ 320
Cabo de rede CAT6 (metro)  | 200 | R$ 2        | R$ 4
Fonte notebook universal   | 6   | R$ 60       | R$ 120
Mouse óptico               | 15  | R$ 25       | R$ 50
Teclado simples            | 10  | R$ 35       | R$ 70

VALOR TOTAL ESTOQUE: Aprox. R$ 31.000

OBSERVAÇÕES (MARIA):
- Estoque baixo de monitores — pedir mais 10 na DigiDistri
- Preços DigiDistri subiram 8% em junho
- Verificar com Intelbras se vai lançar novo modelo de switch
"""
    write_text(os.path.join(ruido_dir, "inventario_estoque.txt"), inventario)

    print(f"  → {sum(1 for _ in os.listdir(os.path.join(base, 'notas-fiscais')) if _.endswith('.csv'))} CSVs de NF")
    print(f"  → Persona NovaTech concluída.")


# ═══════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════

def main():
    print("=" * 60)
    print("GERADOR DE DADOS DE TESTE — PERSONAS BLU")
    print("=" * 60)
    print(f"Data base: {TODAY.isoformat()}")

    gerar_carolina()
    gerar_lucia()
    gerar_novatech()

    print(f"\n{'='*60}")
    print("GERAÇÃO CONCLUÍDA!")
    print(f"{'='*60}")
    print(f"\nEstrutura gerada em: {PERSONAS_DIR}")
    print("\nResumo:")
    for slug in ["carolina-design", "lucia-buffet", "novatech-ti"]:
        nf_path = os.path.join(PERSONAS_DIR, slug, "notas-fiscais")
        csv_count = len([f for f in os.listdir(nf_path) if f.endswith('.csv')]) if os.path.exists(nf_path) else 0
        txt_count = len([f for f in os.listdir(os.path.join(PERSONAS_DIR, slug, "documentos-ruido"))
                         if f.endswith('.txt')])
        xlsx_count = len([f for f in os.listdir(os.path.join(PERSONAS_DIR, slug, "planilhas"))
                          if f.endswith('.xlsx')])
        print(f"  {slug}: {csv_count} CSVs NF, {xlsx_count} planilhas, {txt_count} documentos ruído")


if __name__ == "__main__":
    main()
